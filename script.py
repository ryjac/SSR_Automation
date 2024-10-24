import os
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
from dotenv import load_dotenv

# Define headers
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.5938.132 Safari/537.36"
}

# Start a session
session = requests.Session()

# Persist session
session.headers.update(headers)

# Globally initialize the workbook and worksheet at the start
wb_results = openpyxl.Workbook()
ws_results = wb_results.active

# Create header row
ws_results.append(
    [
        "Subscriber Name",
        "Subscriber ID",
        "Birth Date",
        "Issue Date",
        "Service Date",
        "Primary Aid Code",
        "Responsible County",
        "Trace Number",
        "Eligibility Message",
    ]
)

# Apply bold font to header row
bold_font = Font(bold=True)
for cell in ws_results["1:1"]:
    cell.font = bold_font


def load_credentials():
    # Load the .env file
    load_dotenv()

    # Access Medi-Cal credentials from the .env file
    username = os.environ.get("LOGIN_USERNAME")
    password = os.environ.get("LOGIN_PASSWORD")

    # Check if username and password are None (not found in .env)
    if username is None or password is None:
        print("Error: .env file not found or missing LOGIN_USERNAME or LOGIN_PASSWORD.")
        exit(
            1
        )  # You can choose to handle the error differently, e.g., raise an exception or exit the script
    else:
        print("Credentials loaded")

    return username, password


def login_to_medi_cal():
    # Get credentials
    username, password = load_credentials()

    # Medi-Cal Login URL
    login_url = "https://secure.medi-cal.ca.gov/mcwebpub/login.aspx"

    # Fetch the login page first
    response = session.get(login_url)

    # Use BeautifulSoup to parse the page
    soup = BeautifulSoup(response.content, "html.parser")

    # Get the login form data
    form_data = get_login_form_data_from_soup(soup, username, password)

    # POST the payload
    login_response = session.post(login_url, data=form_data)

    # Check the response to see if the login was successful
    # Update this check based on the expected behavior of a successful login
    if "Single Subscriber" in login_response.text:
        print("Medi-Cal login successful")

        next_page_url = "https://secure.medi-cal.ca.gov/eligibility/Eligibility.aspx"
        response_next_page = session.get(next_page_url)

        if "Subscriber ID" in response_next_page.text:
            print("Navigated to SSR Form")
            print(f"--------------------------------\n")
            submit_ssr()

    else:
        print("Medi-Cal login failed")


def get_login_form_data_from_soup(soup, username, password):
    # Extract hidden fields
    viewstate = soup.find("input", {"name": "__VIEWSTATE"})["value"]
    viewstategenerator = soup.find("input", {"name": "__VIEWSTATEGENERATOR"})["value"]
    eventvalidation = soup.find("input", {"name": "__EVENTVALIDATION"})["value"]

    # Create the payload using the extracted values and values from the entry
    login_data = {
        "__VIEWSTATE": viewstate,
        "__VIEWSTATEGENERATOR": viewstategenerator,
        "__EVENTVALIDATION": eventvalidation,
        "ctl00$MainContent$txtUserID": username,
        "ctl00$MainContent$txtPassword": password,
        "ctl00$MainContent$btnSubmit": "Login",
    }

    return login_data


def get_next_entry(ws, current_row):
    """
    Get the next set of details from the file.
    Returns a dictionary with the details or None if the end of the file is reached.
    """

    recip_id = ws.cell(row=current_row, column=1).value
    if not recip_id:  # end of worksheet reached
        return None

    recip_dob = ws.cell(row=current_row, column=2).value
    recip_doi = ws.cell(row=current_row, column=3).value
    recip_dos = ws.cell(row=current_row, column=4).value

    return {
        "recip_id": recip_id,
        "recip_dob": recip_dob,
        "recip_doi": recip_doi,
        "recip_dos": recip_dos,
    }


def get_ssr_form_data_from_soup(soup, entry):
    # Extract hidden fields
    viewstate = soup.find("input", {"name": "__VIEWSTATE"})["value"]
    viewstategenerator = soup.find("input", {"name": "__VIEWSTATEGENERATOR"})["value"]
    eventvalidation = soup.find("input", {"name": "__EVENTVALIDATION"})["value"]

    # Create the payload using the extracted values and values from the entry
    eligibility_data = {
        "__VIEWSTATE": viewstate,
        "__VIEWSTATEGENERATOR": viewstategenerator,
        "__EVENTVALIDATION": eventvalidation,
        "ctl00$MainContent$RecipID": entry["recip_id"],
        "ctl00$MainContent$RecipDOB": entry["recip_dob"],
        "ctl00$MainContent$RecipDOI": entry["recip_doi"],
        "ctl00$MainContent$RecipDOS": entry["recip_dos"],
        "__ASYNCPOST": "true",
        "ctl00$MainContent$Submit": "Submit",
    }

    return eligibility_data


def post_eligibility_data(data):
    eligibility_page_url = "https://secure.medi-cal.ca.gov/eligibility/Eligibility.aspx"
    response_eligibility = session.post(eligibility_page_url, data=data)
    return response_eligibility


def parse_ssr_results(response):
    # Parse results
    soup = BeautifulSoup(response.text, "html.parser")

    # Extracting the data with error handling

    eligibility_message_elem = soup.find("span", {"id": "MainContent_lblMessages"})
    eligibility_message = (
        eligibility_message_elem.text if eligibility_message_elem else "Not Found"
    )

    subscriber_name_elem = soup.find("div", {"id": "MainContent_trname"})
    subscriber_name = (
        subscriber_name_elem.text.replace("Subscriber Name: ", "").strip()
        if subscriber_name_elem
        else "Not Found"
    )

    subscriber_id_elem = soup.find("div", {"id": "MainContent_trssntrue"})
    subscriber_id = (
        subscriber_id_elem.text.replace("Subscriber ID: ", "").strip()
        if subscriber_id_elem
        else "Not Found"
    )

    birth_date_b_tag = soup.find("b", string="Subscriber Birth Date: ")
    if birth_date_b_tag:
        birth_date = birth_date_b_tag.next_sibling.strip()
    else:
        birth_date = "Not Found"

    issue_date_b_tag = soup.find("b", string="Issue Date: ")
    if issue_date_b_tag:
        issue_date = issue_date_b_tag.next_sibling.strip()
    else:
        issue_date = "Not found"

    primary_aid_code_b_tag = soup.find("b", string="Primary Aid Code: ")
    if primary_aid_code_b_tag:
        primary_aid_code = primary_aid_code_b_tag.next_sibling.strip()
    else:
        primary_aid_code = "Not found"

    responsible_county_b_tag = soup.find("b", string="Responsible County: ")
    if responsible_county_b_tag:
        responsible_county = responsible_county_b_tag.next_sibling.strip()
    else:
        responsible_county = "Not found"

    service_date_b_tag = soup.find("b", string="Service Date: ")
    if service_date_b_tag:
        service_date = service_date_b_tag.next_sibling.strip()
    else:
        service_date = "Not found"

    trace_number_elem = soup.find("span", {"id": "MainContent_lblEVCNo"})
    trace_number = trace_number_elem.text if trace_number_elem else "Not Found"

    return [
        subscriber_name,
        subscriber_id,
        birth_date,
        issue_date,
        service_date,
        primary_aid_code,
        responsible_county,
        trace_number,
        eligibility_message,
    ]


def append_results_to_excel(parsed_data):
    global wb_results
    global ws_results

    # Append the parsed data to the worksheet and save
    ws_results.append(parsed_data)
    print(f"Adding SSR Result to spreadsheet\n")


def get_total_entries(ws):
    total = 0
    for row in ws.iter_rows(
        min_row=2, min_col=1, max_col=1
    ):  # Assuming "recip_id" is in the first column
        if row[0].value:
            total += 1
    return total


def submit_ssr():
    global wb_results
    global ws_results

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook("subscriber_list.xlsx")
    ws = wb.active

    current_row = 2  # assumes first row is header
    total_entries = get_total_entries(ws)

    while True:
        entry = get_next_entry(ws, current_row)
        if entry is None:
            break  # End of worksheet reached, exit loop

        print(f"~~ Running entry ({current_row - 1} of {total_entries}) ~~")

        eligibility_page_url = (
            "https://secure.medi-cal.ca.gov/eligibility/Eligibility.aspx"
        )
        response_eligibility = session.get(eligibility_page_url)
        soup_eligibility = BeautifulSoup(response_eligibility.content, "html.parser")

        form_data = get_ssr_form_data_from_soup(soup_eligibility, entry)
        post_eligibility_data(form_data)

        print("SSR Form submit successful")

        results_page_url = "https://secure.medi-cal.ca.gov/Eligibility/EligResp.aspx"
        response = session.get(results_page_url)

        if "Single Subscriber Response" in response.text:
            print("Obtained SSR Result")
            parsed_data = parse_ssr_results(response)
            append_results_to_excel(parsed_data)
            current_row += 1  # Move to the next row for the next set of details
        else:
            print("Failed to retrieve SSR result")
            break  # Break the loop if SSR form submit fails

    wb_results.save("Results.xlsx")
    print("--------------------------------")
    print("Results saved in 'Results.xlsx'")


if __name__ == "__main__":
    login_to_medi_cal()
